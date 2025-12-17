#!/usr/bin/env python3
"""
author: Yasas Wijesekara (yasas.wijesekara@uni-greifswald.de)

Parses the ICTV datasheet and downloads virus genomes directly
from NCBI in GenBank (gb) format.
"""

from pathlib import Path
import re
import requests
import openpyxl
import click

from vcat.color_logger import logger


NCBI_EFETCH_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
NCBI_LINK_PATTERN = r"https://www\.ncbi\.nlm\.nih\.gov/nuccore/.+"


@click.command(context_settings=dict(show_default=True))
@click.option(
    "-i",
    "--input",
    "infile",
    type=click.Path(path_type=Path, exists=True, dir_okay=False),
    required=True,
    help="ICTV VMR Excel file (.xlsx).",
)
@click.option(
    "-o",
    "--outdir",
    type=click.Path(path_type=Path, file_okay=False),
    required=True,
    help="Output directory to store downloaded GenBank files.",
)
@click.option(
    "--batch-size",
    default=200,
    type=int,
    help="Number of accessions per NCBI request.",
)
def main(infile: Path, outdir: Path, batch_size: int) -> None:
    outdir = outdir.resolve()
    outdir.mkdir(parents=True, exist_ok=True)

    logger.info("Loading VMR metadata: %s", infile)

    wb = openpyxl.load_workbook(infile)
    sheets = wb.sheetnames
    ws = wb[sheets[1]]

    headers = [cell.value for cell in ws[1]]

    dlink = "Accessions Link"
    virus_names = "Virus name(s)"

    all_ids: list[str] = []

    id_list_path = outdir / "ID.list"
    with open(id_list_path, "w") as fh:
        for row in range(2, ws.max_row + 1):
            names = ws.cell(row=row, column=headers.index(virus_names) + 1)
            name_val = "" if names.value is None else str(names.value)

            # skip gene transfer agents
            if "gene transfer agent" in name_val.lower():
                continue

            cell = ws.cell(row=row, column=headers.index(dlink) + 1)
            if not cell.value:
                continue

            # Extract URL inside quotes
            try:
                link = name_val.split('"')[1]
            except IndexError:
                continue

            if not re.search(NCBI_LINK_PATTERN, link):
                continue

            ids = link.split("/")[-1]
            fh.write(ids + "\n")
            all_ids.extend(ids.split("(")[0].split(","))

    logger.info("Found %d accessions. Downloading from NCBI…", len(all_ids))

    for i in range(0, len(all_ids), batch_size):
        params = {
            "db": "nuccore",
            "id": ",".join(all_ids[i : i + batch_size]),
            "rettype": "gb",
            "retmode": "text",
        }

        response = requests.get(NCBI_EFETCH_URL, params=params)

        if response.status_code == 200:
            outfile = outdir / f"sequences_{i}_{i + batch_size}.gb"
            outfile.write_text(response.text)
            logger.info("Downloaded: %s", outfile.name)
        else:
            logger.error(
                "Failed batch %d–%d (HTTP %d)",
                i,
                i + batch_size,
                response.status_code,
            )

    # Sentinel file for downstream workflows
    (outdir / "download_complete").touch()
    logger.info("Download complete.")


if __name__ == "__main__":
    main()
